from flask_restx import Namespace, fields
from app.utils import get_logger, auth, build_ret, curr_date, conn_db
from app.modules import TaskStatus, ErrorMsg
from . import ARLResource, get_arl_parser, base_query_fields
from app import celerytask
import bson

ns = Namespace('icp', description="ICP 备案查询")
logger = get_logger()

# 基础搜索字段
base_search_icp_task_fields = {
    'name': fields.String(required=False, description="任务名"),
    'target': fields.String(description="公司名称"),
    'status': fields.String(description="任务状态"),
    'query_type': fields.String(description="查询类型"),
    '_id': fields.String(description="任务ID")
}
base_search_icp_task_fields.update(base_query_fields)
search_icp_task_fields = ns.model('SearchIcpTask', base_search_icp_task_fields)

# 新建任务参数
add_icp_task_fields = ns.model('AddIcpTask', {
    'name': fields.String(required=True, example="腾讯备案查询", description="任务名"),
    'target': fields.String(required=True, example="深圳市腾讯计算机系统有限公司", description="目标公司"),
    'query_type': fields.List(fields.String, required=True, example=["web", "app"], description="查询类型列表(web/app/mapp/kapp)")
})

add_tyc_task_fields = ns.model('AddTycTask', {
    'name': fields.String(required=True, example="腾讯企业查询", description="任务名"),
    'gid': fields.String(required=True, example="25174642", description="公司 ID(TYC)"),
    'depth': fields.Integer(required=False, default=1, description="递归查询层数"),
    'query_type': fields.List(fields.String, required=True, example=["invest", "trademark", "web", "app", "mapp", "wechat", "weibo"], description="查询类型列表")
})

@ns.route('/task')
class IcpTask(ARLResource):
    parser = get_arl_parser(search_icp_task_fields, location='args')

    @auth
    @ns.expect(parser)
    def get(self):
        """
        获取 ICP 任务列表
        """
        args = self.parser.parse_args()
        data = self.build_data(args=args, collection='icp_task')
        return data

    @auth
    @ns.expect(add_icp_task_fields)
    def post(self):
        """
        新建 ICP 查询任务
        """
        args = self.parse_args(add_icp_task_fields)
        name = args.get('name')
        target = args.get('target')
        query_type = args.get('query_type')

        task_data = {
            "name": name,
            "target": target,
            "query_type": query_type,
            "status": TaskStatus.WAITING,
            "start_time": curr_date(),
            "end_time": "-",
            "statistic": {
                "asset_cnt": 0
            }
        }

        # 存入数据库
        conn = conn_db('icp_task')
        result = conn.insert_one(task_data)
        task_id = str(result.inserted_id)

        # 异步调度 celery
        options = {
            "task_id": task_id,
            "target": target,
            "query_type": query_type
        }

        # 将任务发给 celery
        options["type"] = "icp"
        import requests
        try:
            res = requests.post("http://osint-service:16181/api/v1/recon/start", json=options, timeout=5)
            res.raise_for_status()
        except Exception as e:
            logger.error(f"Failed to trigger osint-service: {e}")
            conn.update_one({"_id": result.inserted_id}, {"$set": {"status": "error", "error_msg": "请求 osint-service 失败或超时"}})

        return build_ret(ErrorMsg.Success, {"task_id": task_id})


@ns.route('/tyc_check')
class TycCheck(ARLResource):
    @auth
    def get(self):
        """
        校验天眼查配置是否有效
        """
        from app.services.tycClient import TycClient
        try:
            client = TycClient()
            success, msg = client.check_token()
            return build_ret(ErrorMsg.Success, {"valid": success, "message": msg})
        except Exception as e:
            return build_ret(ErrorMsg.Success, {"valid": False, "message": f"校验异常: {e}"})


@ns.route('/tyc_task')
class TycTask(ARLResource):
    @auth
    @ns.expect(add_tyc_task_fields)
    def post(self):
        """
        新建企业资产查询任务(TYC)
        """
        args = self.parse_args(add_tyc_task_fields)
        name = args.get('name')
        gid = args.get('gid')
        depth = args.get('depth', 1)
        query_type = args.get('query_type')

        # 校验天眼查配置是否有效
        from app.services.tycClient import TycClient
        try:
            client = TycClient()
            success, msg = client.check_token()
            if not success:
                return build_ret(msg, {})
        except Exception as e:
            return build_ret(f"校验天眼查配置异常: {e}", {})

        task_data = {

            "name": name,
            "target": f"TYC_{gid}",
            "gid": gid,
            "depth": depth,
            "query_type": query_type,
            "task_type": "tyc",  # 用于前端区分
            "status": TaskStatus.WAITING,
            "start_time": curr_date(),
            "end_time": "-",
            "statistic": {
                "asset_cnt": 0
            }
        }

        conn = conn_db('icp_task')
        result = conn.insert_one(task_data)
        task_id = str(result.inserted_id)

        options = {
            "task_id": task_id,
            "gid": gid,
            "depth": depth,
            "query_type": query_type
        }

        options["type"] = "tyc"
        options["tyc_id"] = client.gid
        options["tyc_token"] = client.token
        import requests
        try:
            res = requests.post("http://osint-service:16181/api/v1/recon/start", json=options, timeout=5)
            res.raise_for_status()
        except Exception as e:
            logger.error(f"Failed to trigger osint-service: {e}")
            conn.update_one({"_id": result.inserted_id}, {"$set": {"status": "error", "error_msg": "请求 osint-service 失败或超时"}})

        return build_ret(ErrorMsg.Success, {"task_id": task_id})


# 资产搜索字段
base_search_icp_asset_fields = {
    'task_id': fields.String(description="任务ID"),
    'unitName': fields.String(description="主办单位名称"),
    'domain': fields.String(description="域名"),
    'mainLicence': fields.String(description="主备案号"),
    'companyName': fields.String(description="TYC主办单位"),
    'ym': fields.String(description="TYC域名"),
    'liscense': fields.String(description="TYC主备案号"),
    'webName': fields.String(description="TYC网站名称"),
    'serviceName': fields.String(description="TYC小程序名称"),
    'serviceFilingNumber': fields.String(description="TYC小程序备案号"),
    'name': fields.String(description="TYC名称(APP/投资/微博)"),
    'classes': fields.String(description="TYC分类"),
    'type': fields.String(description="TYC应用类型"),
    'legalPersonName': fields.String(description="TYC法定代表人"),
    'title': fields.String(description="TYC公众号名称"),
    'publicNum': fields.String(description="TYC公众号"),
    'tmName': fields.String(description="TYC商标名称"),
    'regNo': fields.String(description="TYC商标注册号"),
    'regCapital_num': fields.Float(description="TYC注册资本数字(等于)"),
    'regCapital_num__ngt': fields.Float(description="TYC注册资本数字(大于)"),
    'regCapital_num__nlt': fields.Float(description="TYC注册资本数字(小于)"),
    'percent_num': fields.Float(description="TYC投资比例数字(等于)"),
    'percent_num__ngt': fields.Float(description="TYC投资比例数字(大于)"),
    'percent_num__nlt': fields.Float(description="TYC投资比例数字(小于)"),
    'query_type': fields.String(description="查询类型"),
    '_id': fields.String(description="资产ID")
}
base_search_icp_asset_fields.update(base_query_fields)
search_icp_asset_fields = ns.model('SearchIcpAsset', base_search_icp_asset_fields)

@ns.route('/asset')
class IcpAsset(ARLResource):
    parser = get_arl_parser(search_icp_asset_fields, location='args')

    @auth
    @ns.expect(parser)
    def get(self):
        """
        获取 ICP 资产列表
        """
        args = self.parser.parse_args()
        data = self.build_data(args=args, collection='icp_asset')
        return data


@ns.route('/stop/<string:task_id>')
class IcpTaskStop(ARLResource):
    @auth
    def get(self, task_id):
        """停止 ICP 任务 (单任务)"""
        try:
            query = {"_id": bson.ObjectId(task_id)}
            update = {"$set": {"status": TaskStatus.STOP, "end_time": curr_date()}}
            conn_db('icp_task').update_one(query, update)
            return build_ret(ErrorMsg.Success, {"task_id": task_id})
        except Exception as e:
            return build_ret(ErrorMsg.Error, {"error": str(e)})


@ns.route('/delete/')
class IcpTaskBatchDelete(ARLResource):
    @auth
    def post(self):
        """批量删除 ICP 任务"""
        try:
            from flask import request as flask_request
            args = flask_request.json or {}
            task_ids = args.get('task_ids', [])
            if not isinstance(task_ids, list):
                return build_ret(ErrorMsg.Error, {"error": "task_ids must be a list"})

            import bson
            for tid in task_ids:
                conn_db('icp_task').delete_one({"_id": bson.ObjectId(tid)})
                conn_db('icp_asset').delete_many({"task_id": tid})

            return build_ret(ErrorMsg.Success, {"msg": f"成功删除 {len(task_ids)} 个任务"})
        except Exception as e:
            return build_ret(ErrorMsg.Error, {"error": str(e)})

@ns.route('/delete/<string:task_id>')
class IcpTaskDelete(ARLResource):
    @auth
    def get(self, task_id):
        """删除 ICP 任务 (单任务)"""
        try:
            # 删任务和资产
            query = {"_id": bson.ObjectId(task_id)}
            conn_db('icp_task').delete_one(query)
            conn_db('icp_asset').delete_many({"task_id": task_id})
            return build_ret(ErrorMsg.Success, {"task_id": task_id})
        except Exception as e:
            return build_ret(ErrorMsg.Error, {"error": str(e)})


@ns.route('/restart/<string:task_id>')
class IcpTaskRestart(ARLResource):
    @auth
    def get(self, task_id):
        """重启 ICP 任务 (单任务)"""
        try:
            task = conn_db('icp_task').find_one({"_id": bson.ObjectId(task_id)})
            if not task:
                return build_ret(ErrorMsg.NotFoundTask, {"task_id": task_id})

            # 克隆旧任务数据
            task.pop('_id', None)
            task['status'] = TaskStatus.WAITING
            task['start_time'] = curr_date()
            task['end_time'] = "-"
            task['statistic'] = {"asset_cnt": 0}
            task['error_msg'] = ""

            # 插入新任务
            result = conn_db('icp_task').insert_one(task)
            new_task_id = str(result.inserted_id)

            # 重新发起
            if task.get("task_type") == "tyc":
                options = {
                    "task_id": new_task_id,
                    "gid": task.get("gid"),
                    "depth": task.get("depth", 1),
                    "query_type": task.get("query_type", ["web"])
                }
                options["type"] = "tyc"
                options["tyc_id"] = task.get("gid")
                
                from app.services.tycClient import TycClient
                client = TycClient()
                options["tyc_token"] = client.token
                
                import requests
                try:
                    res = requests.post("http://osint-service:16181/api/v1/recon/start", json=options, timeout=5)
                    res.raise_for_status()
                except Exception as e:
                    logger.error(f"Failed to trigger osint-service: {e}")
                    conn_db('icp_task').update_one({"_id": result.inserted_id}, {"$set": {"status": "error", "error_msg": "请求 osint-service 失败或超时"}})
            else:
                options = {
                    "task_id": new_task_id,
                    "target": task.get("target"),
                    "query_type": task.get("query_type", ["web"])
                }
                options["type"] = "icp"
                import requests
                try:
                    res = requests.post("http://osint-service:16181/api/v1/recon/start", json=options, timeout=5)
                    res.raise_for_status()
                except Exception as e:
                    logger.error(f"Failed to trigger osint-service: {e}")
                    conn_db('icp_task').update_one({"_id": result.inserted_id}, {"$set": {"status": "error", "error_msg": "请求 osint-service 失败或超时"}})

            return build_ret(ErrorMsg.Success, {"task_id": new_task_id})
        except Exception as e:
            return build_ret(ErrorMsg.Error, {"error": str(e)})


from flask import make_response
import io
import csv

@ns.route('/export/<string:task_id>')
class IcpTaskExport(ARLResource):
    @auth
    def get(self, task_id):
        """导出 ICP/TYC 资产 (多表单 Excel)"""
        try:
            import openpyxl
            from io import BytesIO

            assets = list(conn_db('icp_asset').find({"task_id": task_id}))
            wb = openpyxl.Workbook()
            wb.remove(wb.active) # 移除默认的 Sheet

            headers = {
                'web': (['主办单位名称', '单位性质', '主备案号', '域名', '网站名称', '服务许可', '更新时间/审核日期'],
                        ['unitName|companyName', 'natureName|companyType', 'mainLicence|liscense', 'domain|ym', 'serviceName|webName', 'serviceLicence', 'updateRecordTime|examineDate']),
                'app': (['APP名称', '主办单位名称', '单位性质/分类', '主备案号/应用类型', 'APP备案号', '审核时间', '简介'],
                        ['name|serviceName', 'unitName', 'natureName|classes', 'mainLicence|type', 'serviceLicence', 'updateRecordTime|examineDate', 'brief']),
                'invest': (['投资公司名称', '法定代表人', '注册资本', '投资比例(%)'], ['name', 'legalPersonName', 'regCapital', 'percent']),
                'trademark': (['商标名称', '注册号', '分类', '状态'], ['tmName', 'regNo', 'intCls', 'status']),
                'wechat': (['公众号名称', '微信号', '简介'], ['title', 'publicNum', 'recommend']),
                'weibo': (['微博名称', '微博链接'], ['name', 'href']),
                'mapp': (['小程序名称/备案号', '审核日期/简介'], ['name|serviceName', 'examineDate|serviceFilingNumber']),
                'kapp': (['快应用名称', '简介'], ['name', 'brief']),
            }

            grouped_assets = {}
            for item in assets:
                qt = item.get('query_type', 'unknown')
                if qt not in grouped_assets:
                    grouped_assets[qt] = []
                # 兼容 mapp 的新老接口字段
                if qt == 'mapp' and not item.get('name'):
                    item['name'] = item.get('serviceName', '')
                grouped_assets[qt].append(item)

            for qt, items in grouped_assets.items():
                if qt in headers:
                    h_labels, h_keys = headers[qt]
                else:
                    h_labels = ['数据名称']
                    h_keys = ['name']

                ws = wb.create_sheet(title=qt)
                ws.append(h_labels)
                for item in items:
                    row = []
                    for k in h_keys:
                        if '|' in k:
                            k1, k2 = k.split('|')
                            val = item.get(k1) if item.get(k1) else item.get(k2, '')
                        else:
                            val = item.get(k, '')
                        row.append(str(val))
                    ws.append(row)

            if len(wb.sheetnames) == 0:
                ws = wb.create_sheet(title='Empty')
                ws.append(['无数据'])

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = f"attachment; filename=export.xlsx"
            response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return response
        except Exception as e:
            return build_ret(ErrorMsg.Error, {"error": str(e)})


@ns.route('/sync/<string:task_id>')
class IcpTaskSync(ARLResource):
    @auth
    def post(self, task_id):
        """将 ICP/TYC 任务中的网站域名同步至资产分组"""
        try:
            from flask import request as flask_request
            args = flask_request.json or {}
            mode = args.get('mode', 'new') # 'new' or 'existing'
            target_name = args.get('target_name', '').strip()
            scope_id = args.get('scope_id')

            task = conn_db('icp_task').find_one({"_id": bson.ObjectId(task_id)})
            if not task:
                return build_ret(ErrorMsg.NotFoundTask, {"task_id": task_id})

            # 查出当前任务下所有 web 查询的资产 (只取网站)
            assets = list(conn_db('icp_asset').find({"task_id": task_id, "query_type": "web"}))
            domains = set()
            for asset in assets:
                d = asset.get('domain') or asset.get('ym')
                if d and isinstance(d, str):
                    domains.add(d.strip())

            if not domains:
                return build_ret(ErrorMsg.Error, {"error": "未发现网站资产，无法同步"})

            if mode == 'existing' and scope_id:
                # 关联已有资产
                scope = conn_db('asset_scope').find_one({"_id": bson.ObjectId(scope_id)})
                if not scope:
                    return build_ret(ErrorMsg.Error, {"error": "指定的资产组不存在"})
                old_array = scope.get("scope_array", [])
                new_array = list(set(old_array) | domains)
                old_domain_array = scope.get("domain_array")
                if old_domain_array is None:
                    old_domain_array = old_array if scope.get("scope_type", "domain") == "domain" else []
                new_domain_array = list(set(old_domain_array) | domains)
                conn_db('asset_scope').update_one(
                    {"_id": scope["_id"]},
                    {"$set": {
                        "scope_array": new_array,
                        "scope": ",".join(new_array),
                        "domain_array": new_domain_array
                    }}
                )
                target_name = scope.get('name')
            else:
                # 新建资产组
                if not target_name:
                    target_name = task.get('target', '未知公司')

                # 检查是否同名，不再限制 scope_type
                scope = conn_db('asset_scope').find_one({"name": target_name})
                if scope:
                    old_array = scope.get("scope_array", [])
                    new_array = list(set(old_array) | domains)
                    old_domain_array = scope.get("domain_array")
                    if old_domain_array is None:
                        old_domain_array = old_array if scope.get("scope_type", "domain") == "domain" else []
                    new_domain_array = list(set(old_domain_array) | domains)
                    conn_db('asset_scope').update_one(
                        {"_id": scope["_id"]},
                        {"$set": {
                            "scope_array": new_array,
                            "scope": ",".join(new_array),
                            "domain_array": new_domain_array
                        }}
                    )
                else:
                    new_array = list(domains)
                    scope_data = {
                        "name": target_name,
                        "scope_type": "mixed",
                        "scope": ",".join(new_array),
                        "scope_array": new_array,
                        "domain_array": new_array,
                        "ip_array": [],
                        "black_scope": "",
                        "black_scope_array": []
                    }
                    conn_db('asset_scope').insert_one(scope_data)

            return build_ret(ErrorMsg.Success, {"msg": f"成功同步了 {len(domains)} 个域名到资产分组 '{target_name}'"})
        except Exception as e:
            return build_ret(ErrorMsg.Error, {"error": str(e)})
